from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import json, io

# ═══════════════════════════════════════════
# SHARED STYLES
# ═══════════════════════════════════════════
DARK='1E3A5F'; TEAL='00A896'; TEAL2='00505A'; WHITE='FFFFFF'
ALT1='EEF6F7'; ALT2='FFFFFF'; RED='C00000'; ORANGE='E07B2A'; GREEN='00704A'
MONTHS=['يناير','فبراير','مارس','أبريل','مايو','يونيو',
        'يوليو','أغسطس','سبتمبر','أكتوبر','نوفمبر','ديسمبر']

def bd(col='B2D8D8'):
    s=Side(border_style='thin',color=col)
    return Border(left=s,right=s,top=s,bottom=s)

def Al(h='center',wrap=False):
    return Alignment(horizontal=h,vertical='center',readingOrder=2,wrapText=wrap)

def C(ws,r,c,val='',bold=False,fg='000000',bg=None,h='center',size=10,wrap=False):
    x=ws.cell(row=r,column=c,value=val)
    x.font=Font(name='Arial',bold=bold,size=size,color=fg)
    if bg: x.fill=PatternFill('solid',start_color=bg)
    x.alignment=Al(h,wrap); x.border=bd(); return x

def MC(ws,r,c,span,val,bold=True,fg=WHITE,bg=DARK,size=13,h='center'):
    ws.merge_cells(start_row=r,start_column=c,end_row=r,end_column=c+span-1)
    x=ws.cell(row=r,column=c,value=val)
    x.font=Font(name='Arial',bold=bold,size=size,color=fg)
    x.fill=PatternFill('solid',start_color=bg)
    x.alignment=Al(h); x.border=bd(WHITE); return x

def RH(ws,r,h): ws.row_dimensions[r].height=h
def CW(ws,c,w): ws.column_dimensions[get_column_letter(c) if isinstance(c,int) else c].width=w

def TH(ws,r,cols,bg=TEAL2):
    for ci,h in enumerate(cols,1):
        x=ws.cell(r,ci,h)
        x.font=Font(name='Arial',bold=True,size=10,color=WHITE)
        x.fill=PatternFill('solid',start_color=bg)
        x.alignment=Al(wrap=True); x.border=bd(WHITE)
    RH(ws,r,28)

def KPI(ws,r,col,label,val,vbg):
    ws.merge_cells(start_row=r,start_column=col,end_row=r,end_column=col+1)
    lc=ws.cell(r,col,label)
    lc.font=Font(name='Arial',bold=True,size=10,color=WHITE)
    lc.fill=PatternFill('solid',start_color=TEAL2)
    lc.alignment=Al('right'); lc.border=bd(WHITE)
    ws.merge_cells(start_row=r,start_column=col+2,end_row=r,end_column=col+3)
    vc=ws.cell(r,col+2,val)
    vc.font=Font(name='Arial',bold=True,size=12,color=WHITE)
    vc.fill=PatternFill('solid',start_color=vbg)
    vc.alignment=Al(); vc.border=bd(WHITE)
    RH(ws,r,26)

# ═══════════════════════════════════════════
# BUILDERS
# ═══════════════════════════════════════════

def add_violations_sheet(wb, recs, sheet_title, year):
    """إضافة شيت مخالفات واحد للـ workbook"""
    ds=wb.create_sheet(sheet_title); ds.sheet_view.rightToLeft=True
    total=len(recs)
    MC(ds,1,1,8,f'{sheet_title}  |  نورة لتعليم القيادة',size=13); RH(ds,1,34)
    MC(ds,2,1,8,f'إجمالي المخالفات: {total}  |  السنة: {year}',bold=False,bg=TEAL,size=10); RH(ds,2,22)
    TH(ds,3,['اسم المدرب/ة','ZONE','الوقت','التاريخ','نوع المخالفة','شرح المخالفة','درجة المخالفة','تم الرصد من قبل'])
    grade_colors={'A':RED,'B':ORANGE,'C':'006B3C'}
    for ri,rec in enumerate(recs):
        r=4+ri; bg=ALT1 if ri%2==0 else ALT2
        C(ds,r,1,rec.get('trainer_name','—'),bold=True,fg=DARK,bg=bg,h='right')
        C(ds,r,2,rec.get('zone',''),bg=bg)
        C(ds,r,3,rec.get('time',''),bg=bg)
        C(ds,r,4,rec.get('date',''),bg=bg)
        C(ds,r,5,rec.get('violation_type',''),bg=bg,h='right',wrap=True)
        C(ds,r,6,rec.get('description',''),bg=bg,h='right',wrap=True)
        grade=rec.get('grade','')
        gc=ds.cell(r,7,grade)
        gc.font=Font(name='Arial',bold=True,size=10,color=grade_colors.get(grade,WHITE))
        gc.fill=PatternFill('solid',start_color=bg); gc.alignment=Al(); gc.border=bd()
        C(ds,r,8,rec.get('observer',''),bg=bg)
        RH(ds,r,20)
    ds.column_dimensions['A'].width=28; ds.column_dimensions['B'].width=8
    ds.column_dimensions['C'].width=10; ds.column_dimensions['D'].width=12
    ds.column_dimensions['E'].width=34; ds.column_dimensions['F'].width=30
    ds.column_dimensions['G'].width=14; ds.column_dimensions['H'].width=20
    ds.freeze_panes='A4'

def add_violations_kpi(wb, all_recs, year):
    """شيت KPIs للمخالفات"""
    ks=wb.active; ks.title='لوحة KPIs'; ks.sheet_view.rightToLeft=True
    female=[r for r in all_recs if r.get('section')=='female']
    male=[r for r in all_recs if r.get('section')=='male']
    total=len(all_recs)
    gA=sum(1 for r in all_recs if r.get('grade')=='A')
    gB=sum(1 for r in all_recs if r.get('grade')=='B')
    gC=sum(1 for r in all_recs if r.get('grade')=='C')
    MC(ks,1,1,10,f'لوحة المخالفات – KPIs  |  نورة لتعليم القيادة',size=14); RH(ks,1,36)
    MC(ks,2,1,10,f'إجمالي السجلات: {total}  |  السنة: {year}',bold=False,bg=TEAL,size=10); RH(ks,2,20)
    MC(ks,4,1,10,'📊 مؤشرات الأداء الرئيسية',bg=TEAL,size=11); RH(ks,4,26)
    KPI(ks,5,1,'إجمالي المخالفات',total,DARK)
    KPI(ks,6,1,'مخالفات A – خطيرة',gA,RED)
    KPI(ks,7,1,'مخالفات B – متوسطة',gB,ORANGE)
    KPI(ks,8,1,'مخالفات C – خفيفة',gC,'006B3C')
    KPI(ks,9,1,'مخالفات القسم النسائي',len(female),TEAL2)
    KPI(ks,10,1,'مخالفات القسم الرجالي',len(male),TEAL2)
    for c,w in enumerate([20,10,10,10,20,10,10,10,10,10],1): CW(ks,c,w)
    # جدول شهري
    MC(ks,12,1,8,'📅 توزيع المخالفات الشهري',bg=TEAL,size=11); RH(ks,12,26)
    TH(ks,13,['الشهر','الإجمالي','مخالفات A','مخالفات B','مخالفات C','نسبة A%','نسبة B%','نسبة C%'])
    for mi,month in enumerate(MONTHS):
        r=14+mi; bg=ALT1 if mi%2==0 else ALT2
        mr=[x for x in all_recs if (x.get('date') or '').startswith(f'{year}-{str(mi+1).zfill(2)}')]
        mt=len(mr); mA=sum(1 for x in mr if x.get('grade')=='A')
        mB=sum(1 for x in mr if x.get('grade')=='B')
        mC=sum(1 for x in mr if x.get('grade')=='C')
        C(ks,r,1,month,bold=True,fg=DARK,bg=bg,h='right')
        C(ks,r,2,mt,bg=bg); C(ks,r,3,mA,fg=RED,bg=bg)
        C(ks,r,4,mB,fg=ORANGE,bg=bg); C(ks,r,5,mC,fg='006B3C',bg=bg)
        C(ks,r,6,f'{mA/mt*100:.1f}%' if mt else '-',bg=bg)
        C(ks,r,7,f'{mB/mt*100:.1f}%' if mt else '-',bg=bg)
        C(ks,r,8,f'{mC/mt*100:.1f}%' if mt else '-',bg=bg)
        RH(ks,r,20)
    for c,w in enumerate([12,10,10,10,10,10,10,10],1): CW(ks,c,w)

def build_violations(recs, section_name, year):
    wb=Workbook()
    female=[r for r in recs if r.get('section')=='female']
    male=[r for r in recs if r.get('section')=='male']
    # شيت KPIs أولاً
    add_violations_kpi(wb, recs, year)
    # شيت النساء
    add_violations_sheet(wb, female, 'مخالفات القسم النسائي', year)
    # شيت الرجال
    add_violations_sheet(wb, male, 'مخالفات القسم الرجالي', year)
    return wb

def build_surveys(recs, year):
    wb=Workbook()
    ws=wb.active; ws.title='بيانات الاستطلاع'; ws.sheet_view.rightToLeft=True
    total=len(recs)
    scores=[((r.get('q1',0) or 0)+(r.get('q2',0) or 0)+(r.get('q3',0) or 0)+(r.get('q4',0) or 0)+(r.get('q5',0) or 0))/25*100 for r in recs]
    avg=sum(scores)/len(scores) if scores else 0
    
    MC(ws,1,1,11,f'تقرير استطلاع رضا العملاء  |  إدارة الجودة  |  نورة لتعليم القيادة',size=13); RH(ws,1,34)
    MC(ws,2,1,11,f'إجمالي الاستطلاعات: {total}  |  متوسط الرضا: {avg:.1f}%  |  السنة: {year}',bold=False,bg=TEAL,size=10); RH(ws,2,22)
    TH(ws,3,['م','التاريخ','الموظف','خدمة العملاء/5','التدريب النظري/5','مبنى العملي/5','التدريب العملي/5','ما بعد التدريب/5','المجموع/25','نسبة الرضا%','ملاحظات'])
    
    for ri,rec in enumerate(recs):
        r=4+ri; bg=ALT1 if ri%2==0 else ALT2
        q=[rec.get(f'q{i}',0) or 0 for i in range(1,6)]
        total_score=sum(q); pct=total_score/25*100
        pct_bg='009975' if pct>=90 else ORANGE if pct>=75 else RED
        C(ws,r,1,ri+1,bg=bg); C(ws,r,2,rec.get('date',''),bg=bg)
        C(ws,r,3,rec.get('observer',''),bold=True,fg=DARK,bg=bg,h='right')
        for qi,qv in enumerate(q,1): C(ws,r,3+qi,qv,bg=bg)
        C(ws,r,9,total_score,bold=True,bg=bg)
        pctc=ws.cell(r,10,f'{pct:.1f}%')
        pctc.font=Font(name='Arial',bold=True,size=10,color=WHITE)
        pctc.fill=PatternFill('solid',start_color=pct_bg); pctc.alignment=Al(); pctc.border=bd()
        C(ws,r,11,rec.get('notes',''),bg=bg,wrap=True); RH(ws,r,20)
    
    for c,w in [(1,5),(2,12),(3,18),(4,12),(5,14),(6,12),(7,14),(8,16),(9,10),(10,11),(11,20)]: CW(ws,c,w)
    ws.freeze_panes='A4'
    
    # لوحة التحليل
    la=wb.create_sheet('لوحة التحليل'); la.sheet_view.rightToLeft=True
    MC(la,1,1,8,f'لوحة تحليل استطلاع رضا العملاء  |  نورة لتعليم القيادة',size=13); RH(la,1,34)
    MC(la,2,1,8,'تحليل تلقائي من البيانات المسجلة',bold=False,bg=TEAL,size=10); RH(la,2,20)
    KPI(la,4,1,'إجمالي الاستطلاعات',total,DARK)
    KPI(la,5,1,'متوسط الرضا العام',f'{avg:.1f}%',TEAL2 if avg>=90 else ORANGE if avg>=75 else RED)
    KPI(la,6,1,'الهدف','90%','006B3C')
    KPI(la,7,1,'الفجوة عن الهدف',f'{90-avg:+.1f}%',RED if avg<90 else '006B3C')
    for c,w in enumerate([20,10,10,10,20,10,10,10],1): CW(la,c,w)
    return wb

def build_facility(recs, year):
    wb=Workbook()
    ws=wb.active; ws.title='سجل الملاحظات'; ws.sheet_view.rightToLeft=True
    total=len(recs)
    MC(ws,1,1,8,f'تقرير ملاحظات المرافق والنظافة  |  إدارة الجودة  |  نورة لتعليم القيادة',size=13); RH(ws,1,34)
    MC(ws,2,1,8,f'إجمالي الملاحظات: {total}  |  السنة: {year}',bold=False,bg=TEAL,size=10); RH(ws,2,22)
    TH(ws,3,['م','التاريخ','الوقت','نوع المشكلة','الموقع','الوصف','الراصد','الحالة'])
    for ri,rec in enumerate(recs):
        r=4+ri; bg=ALT1 if ri%2==0 else ALT2
        C(ws,r,1,ri+1,bg=bg); C(ws,r,2,rec.get('date',''),bg=bg)
        C(ws,r,3,rec.get('time',''),bg=bg)
        C(ws,r,4,rec.get('problem_type',''),bold=True,fg=DARK,bg=bg,h='right')
        C(ws,r,5,rec.get('location',''),bg=bg,h='right')
        C(ws,r,6,rec.get('description',''),bg=bg,h='right',wrap=True)
        C(ws,r,7,rec.get('observer',''),bg=bg); C(ws,r,8,'',bg=bg); RH(ws,r,20)
    for c,w in [(1,5),(2,12),(3,10),(4,22),(5,16),(6,34),(7,18),(8,16)]: CW(ws,c,w)
    ws.freeze_panes='A4'
    return wb

def build_building(recs, year):
    wb=Workbook()
    ws=wb.active; ws.title='سجل التقارير'; ws.sheet_view.rightToLeft=True
    total=len(recs)
    MC(ws,1,1,9,f'تقرير المبنى الرئيسي  |  إدارة الجودة  |  نورة لتعليم القيادة',size=13); RH(ws,1,34)
    MC(ws,2,1,9,f'إجمالي التقارير: {total}  |  السنة: {year}',bold=False,bg=TEAL,size=10); RH(ws,2,22)
    TH(ws,3,['م','التاريخ','الوقت','ازدحام الاستقبال','موظفات الاستقبال','وقت الانتظار','ازدحام القاعات','ملاحظات عامة','الراصد'])
    crowd_colors={'شديد':RED,'متوسط':ORANGE,'خفيف':'006B3C'}
    for ri,rec in enumerate(recs):
        r=4+ri; bg=ALT1 if ri%2==0 else ALT2
        C(ws,r,1,ri+1,bg=bg); C(ws,r,2,rec.get('date',''),bg=bg); C(ws,r,3,rec.get('time',''),bg=bg)
        rc=rec.get('recep_crowd','')
        rcc=ws.cell(r,4,rc); rcc.font=Font(name='Arial',bold=bool(rc),size=10,color=crowd_colors.get(rc,'000000'))
        rcc.fill=PatternFill('solid',start_color=bg); rcc.alignment=Al(); rcc.border=bd()
        C(ws,r,5,rec.get('recep_staff',''),bg=bg)
        C(ws,r,6,rec.get('wait_time',''),bg=bg)
        ec=rec.get('exam_crowd','')
        ecc=ws.cell(r,7,ec); ecc.font=Font(name='Arial',bold=bool(ec),size=10,color=crowd_colors.get(ec,'000000'))
        ecc.fill=PatternFill('solid',start_color=bg); ecc.alignment=Al(); ecc.border=bd()
        C(ws,r,8,rec.get('general_notes',''),bg=bg,h='right',wrap=True)
        C(ws,r,9,rec.get('observer',''),bg=bg); RH(ws,r,20)
    for c,w in [(1,5),(2,12),(3,10),(4,16),(5,16),(6,14),(7,14),(8,34),(9,18)]: CW(ws,c,w)
    ws.freeze_panes='A4'
    return wb

def build_attendance(recs, month_num, year):
    import datetime
    wb=Workbook()
    ws=wb.active
    month_name=MONTHS[month_num-1]
    ws.title='كشف الحضور'; ws.sheet_view.rightToLeft=True
    
    # أيام العمل
    work_days=[]
    d=datetime.date(year,month_num,1)
    while d.month==month_num:
        if d.weekday() in (6,0,1,2,3): work_days.append(d)
        d+=datetime.timedelta(days=1)
    
    STAFF=['شهد المقرن','سمية الخضيري','امجاد ابانمي','احمد الحماد','ابراهيم الحمد','احمد سحاب','محمد محمد الجمعان']
    SECS={'شهد المقرن':'نسائي','سمية الخضيري':'نسائي','امجاد ابانمي':'نسائي','احمد الحماد':'رجالي','ابراهيم الحمد':'رجالي','احمد سحاب':'رجالي','محمد محمد الجمعان':'إدارة'}
    DN={6:'أحد',0:'اثنين',1:'ثلاثاء',2:'أربعاء',3:'خميس'}
    SYM={'غياب':'❌','إجازة':'🔵','استئذان':'🟡','خروج مبكر':'🔶','تأخير':'⏰'}
    SBG={'✅':'E8F8F2','❌':'FDECEA','🔵':'EAF2FB','🟡':'FFF8E1','🔶':'FFF3E0','⏰':'FFF3E0'}
    SFG={'✅':'00A896','❌':'C00000','🔵':'5299E0','🟡':'E0A800','🔶':'E07B2A','⏰':'E07B2A'}
    
    total_cols=3+len(work_days)+5
    MC(ws,1,1,total_cols,f'كشف حضور وانصراف الموظفين  ·  {month_name} {year}  ·  نورة لتعليم القيادة',size=14); RH(ws,1,36)
    MC(ws,2,1,total_cols,f'أيام العمل: الأحد — الخميس  |  إجمالي أيام العمل: {len(work_days)} يوم',bold=False,bg=TEAL,size=10); RH(ws,2,22)
    
    # رؤوس
    for ci,h in enumerate(['م','اسم الموظف','القسم'],1):
        x=ws.cell(3,ci,h); x.font=Font(name='Arial',bold=True,size=10,color=WHITE)
        x.fill=PatternFill('solid',start_color=DARK); x.alignment=Al(wrap=True); x.border=bd(WHITE)
    for di,day in enumerate(work_days):
        col=4+di; lbl=f"{day.strftime('%d/%m')}\n{DN[day.weekday()]}"
        x=ws.cell(3,col,lbl); x.font=Font(name='Arial',bold=True,size=9,color=WHITE)
        x.fill=PatternFill('solid',start_color=DARK); x.alignment=Al(wrap=True); x.border=bd(WHITE); CW(ws,col,8)
    for si,sh in enumerate(['الحضور','الغياب','استئذان','إجازة','نسبة الحضور%']):
        col=4+len(work_days)+si
        x=ws.cell(3,col,sh); x.font=Font(name='Arial',bold=True,size=9,color=WHITE)
        x.fill=PatternFill('solid',start_color=DARK); x.alignment=Al(wrap=True); x.border=bd(WHITE); CW(ws,col,11)
    RH(ws,3,30); CW(ws,1,5); CW(ws,2,22); CW(ws,3,10)
    
    # بيانات
    monthStr=f'{year}-{str(month_num).zfill(2)}'
    monthRecs=[r for r in recs if (r.get('date') or '').startswith(monthStr)]
    
    for ri,staff in enumerate(STAFF):
        r=4+ri; bg=ALT1 if ri%2==0 else ALT2
        C(ws,r,1,ri+1,bg=bg); C(ws,r,2,staff,bold=True,fg=DARK,bg=bg,h='right')
        C(ws,r,3,SECS.get(staff,'—'),bg=bg)
        p=ab=pm=lv=0
        for di,day in enumerate(work_days):
            col=4+di; dateStr=day.isoformat()
            rec=next((x for x in monthRecs if x.get('staff')==staff and x.get('date')==dateStr),None)
            if not rec: sym='✅'; p+=1
            else:
                sym=SYM.get(rec.get('status',''),'✅')
                st=rec.get('status','')
                if st=='غياب': ab+=1
                elif st=='إجازة': lv+=1
                elif st in ('استئذان','خروج مبكر','تأخير'): pm+=1
                else: p+=1
            x=ws.cell(r,col,sym)
            x.font=Font(name='Arial',bold=(sym=='✅'),size=10,color=SFG.get(sym,'000000'))
            x.fill=PatternFill('solid',start_color=SBG.get(sym,bg)); x.alignment=Al(); x.border=bd()
        sc=4+len(work_days); pct=round(p/len(work_days)*100,1)
        pct_bg='009975' if pct>=90 else ORANGE if pct>=75 else RED
        for off,val,fg in [(0,p,'00A896'),(1,ab,RED),(2,pm,ORANGE),(3,lv,'5299E0')]:
            C(ws,r,sc+off,val,fg=fg,bg=bg)
        pctc=ws.cell(r,sc+4,f'{pct}%')
        pctc.font=Font(name='Arial',bold=True,size=10,color=WHITE)
        pctc.fill=PatternFill('solid',start_color=pct_bg); pctc.alignment=Al(); pctc.border=bd()
        RH(ws,r,20)
    
    # صف الإجماليات
    tr=4+len(STAFF)
    MC(ws,tr,1,3,'الإجمالي',size=10); RH(ws,tr,22)
    for di in range(len(work_days)):
        col=4+di; cl=get_column_letter(col)
        x=ws.cell(tr,col,f'=COUNTIF({cl}4:{cl}{tr-1},"✅")')
        x.font=Font(name='Arial',bold=True,size=9,color=WHITE)
        x.fill=PatternFill('solid',start_color=TEAL); x.alignment=Al(); x.border=bd(WHITE)
    sc=4+len(work_days)
    for off in range(4):
        col=sc+off; cl=get_column_letter(col)
        x=ws.cell(tr,col,f'=SUM({cl}4:{cl}{tr-1})')
        x.font=Font(name='Arial',bold=True,size=10,color=WHITE)
        x.fill=PatternFill('solid',start_color=DARK); x.alignment=Al(); x.border=bd(WHITE)
    p_cl=get_column_letter(sc)
    pctc=ws.cell(tr,sc+4,f'=ROUND({p_cl}{tr}/({len(work_days)}*{len(STAFF)})*100,1)&"%"')
    pctc.font=Font(name='Arial',bold=True,size=10,color=WHITE)
    pctc.fill=PatternFill('solid',start_color=DARK); pctc.alignment=Al(); pctc.border=bd(WHITE)
    ws.freeze_panes='D4'
    
    # شيت التحليل
    ws2=wb.create_sheet('تحليل شهري'); ws2.sheet_view.rightToLeft=True
    MC(ws2,1,1,7,f'تحليل شهري للحضور – {month_name} {year}  ·  نورة لتعليم القيادة',size=13); RH(ws2,1,34)
    for c,w in [(1,24),(2,12),(3,12),(4,12),(5,15),(6,12),(7,14)]: CW(ws2,c,w)
    MC(ws2,3,1,7,'ملخص حضور الموظفين',bg=TEAL,size=11); RH(ws2,3,26)
    TH(ws2,4,['الموظف','القسم','حضور','غياب','استئذان','إجازة','نسبة الحضور%'])
    for ri,staff in enumerate(STAFF):
        r=5+ri; bg=ALT1 if ri%2==0 else ALT2
        staffRecs=[x for x in monthRecs if x.get('staff')==staff]
        p=sum(1 for d in work_days if not next((x for x in staffRecs if x.get('date')==d.isoformat()),None))
        ab=sum(1 for x in staffRecs if x.get('status')=='غياب')
        pm=sum(1 for x in staffRecs if x.get('status') in ('استئذان','خروج مبكر','تأخير'))
        lv=sum(1 for x in staffRecs if x.get('status')=='إجازة')
        pct=round(p/len(work_days)*100,1)
        pct_bg='009975' if pct>=90 else ORANGE if pct>=75 else RED
        for ci,val in enumerate([staff,SECS.get(staff,'—'),p,ab,pm,lv],1):
            C(ws2,r,ci,val,bold=(ci==1),fg=DARK if ci==1 else '000000',bg=bg,h='right' if ci==1 else 'center')
        pctc=ws2.cell(r,7,f'{pct}%')
        pctc.font=Font(name='Arial',bold=True,size=10,color=WHITE)
        pctc.fill=PatternFill('solid',start_color=pct_bg); pctc.alignment=Al(); pctc.border=bd()
        RH(ws2,r,20)
    return wb

def build_incentives(recs, year):
    wb=Workbook()
    ws=wb.active; ws.title='قائمة الحوافز'; ws.sheet_view.rightToLeft=True
    total=len(recs)
    MC(ws,1,1,8,f'تقرير الحوافز الشهرية  |  إدارة الجودة  |  نورة لتعليم القيادة',size=13); RH(ws,1,34)
    MC(ws,2,1,8,f'إجمالي السجلات: {total}  |  السنة: {year}',bold=False,bg=TEAL,size=10); RH(ws,2,22)
    TH(ws,3,['الشهر','السنة','اسم الموظف','القسم','النقاط','المخالفات','الأهلية','ملاحظات'])
    for ri,rec in enumerate(recs):
        r=4+ri; bg=ALT1 if ri%2==0 else ALT2
        score=rec.get('score') or 0; viols=rec.get('violations_count') or 0
        eligible='✅ مؤهل' if score>=80 and viols==0 else '❌ غير مؤهل'
        elig_fg='006B3C' if '✅' in eligible else RED
        C(ws,r,1,rec.get('month',''),bg=bg); C(ws,r,2,rec.get('year',''),bg=bg)
        C(ws,r,3,rec.get('staff_name',''),bold=True,fg=DARK,bg=bg,h='right')
        C(ws,r,4,rec.get('section',''),bg=bg); C(ws,r,5,score,bg=bg); C(ws,r,6,viols,bg=bg)
        ec=ws.cell(r,7,eligible)
        ec.font=Font(name='Arial',bold=True,size=10,color=elig_fg)
        ec.fill=PatternFill('solid',start_color=bg); ec.alignment=Al(); ec.border=bd()
        C(ws,r,8,rec.get('notes',''),bg=bg,h='right',wrap=True); RH(ws,r,20)
    for c,w in [(1,12),(2,8),(3,22),(4,12),(5,10),(6,12),(7,14),(8,24)]: CW(ws,c,w)
    ws.freeze_panes='A4'
    return wb

def build_ops(recs, year):
    wb=Workbook()
    ws=wb.active; ws.title='بيانات المدربين'; ws.sheet_view.rightToLeft=True
    total=len(recs)
    MC(ws,1,1,9,f'تقرير كفاءة التشغيل  |  إدارة الجودة  |  نورة لتعليم القيادة',size=13); RH(ws,1,34)
    MC(ws,2,1,9,f'إجمالي السجلات: {total}  |  السنة: {year}',bold=False,bg=TEAL,size=10); RH(ws,2,22)
    TH(ws,3,['اسم المدرب','الشهر','الزون','القسم','الساعات المتاحة','الحصص','المتدربين','الحصص غير المستغلة','نسبة الكفاءة%'])
    effs=[r.get('efficiency_pct') or 0 for r in recs if r.get('efficiency_pct')]
    avg_eff=sum(effs)/len(effs) if effs else 0
    for ri,rec in enumerate(recs):
        r=4+ri; bg=ALT1 if ri%2==0 else ALT2
        eff=rec.get('efficiency_pct') or 0
        eff_bg='009975' if eff>=85 else ORANGE if eff>=70 else RED
        C(ws,r,1,rec.get('trainer_name',''),bold=True,fg=DARK,bg=bg,h='right')
        C(ws,r,2,MONTHS[int(rec.get('report_month',1))-1] if rec.get('report_month') else '',bg=bg)
        C(ws,r,3,rec.get('zone',''),bg=bg); C(ws,r,4,rec.get('section',''),bg=bg)
        C(ws,r,5,rec.get('available_hours',''),bg=bg); C(ws,r,6,rec.get('sessions',''),bg=bg)
        C(ws,r,7,rec.get('trainees',''),bg=bg)
        unused=(rec.get('sessions') or 0)-(rec.get('trainees') or 0)
        C(ws,r,8,max(0,unused) if unused else '',bg=bg)
        effc=ws.cell(r,9,f'{eff:.1f}%' if eff else '-')
        effc.font=Font(name='Arial',bold=True,size=10,color=WHITE)
        effc.fill=PatternFill('solid',start_color=eff_bg); effc.alignment=Al(); effc.border=bd()
        RH(ws,r,20)
    for c,w in [(1,28),(2,10),(3,8),(4,10),(5,14),(6,10),(7,12),(8,16),(9,13)]: CW(ws,c,w)
    ws.freeze_panes='A4'
    
    # KPIs
    ks=wb.create_sheet('لوحة KPIs'); ks.sheet_view.rightToLeft=True
    MC(ks,1,1,8,f'لوحة كفاءة التشغيل  |  نورة لتعليم القيادة',size=13); RH(ks,1,34)
    KPI(ks,3,1,'إجمالي المدربين',total,DARK)
    KPI(ks,4,1,'متوسط الكفاءة',f'{avg_eff:.1f}%','009975' if avg_eff>=85 else ORANGE if avg_eff>=70 else RED)
    KPI(ks,5,1,'إجمالي الحصص',sum(r.get('sessions') or 0 for r in recs),TEAL2)
    KPI(ks,6,1,'إجمالي المتدربين',sum(r.get('trainees') or 0 for r in recs),TEAL2)
    for c,w in enumerate([20,10,10,10,20,10,10,10],1): CW(ks,c,w)
    return wb

def build_sales(recs, year):
    wb=Workbook()
    ws=wb.active; ws.title='بيانات الاسترداد'; ws.sheet_view.rightToLeft=True
    total=len(recs)
    MC(ws,1,1,8,f'نموذج متابعة عملاء الاسترداد  |  إدارة الجودة  |  نورة لتعليم القيادة',size=13); RH(ws,1,34)
    MC(ws,2,1,8,f'إجمالي العملاء: {total}  |  السنة: {year}',bold=False,bg=TEAL,size=10); RH(ws,2,22)
    TH(ws,3,['م','اسم العميل','رسوم البرنامج','تاريخ الطلب','شرح رغبة العميل','ملاحظات المبيعات','نوع الطلب','الأسبوع'])
    type_colors={'استرداد':RED,'استكمال':'006B3C','تأجيل':ORANGE}
    for ri,rec in enumerate(recs):
        r=4+ri; bg=ALT1 if ri%2==0 else ALT2
        t=rec.get('request_type','')
        C(ws,r,1,ri+1,bg=bg); C(ws,r,2,rec.get('client_name',''),bold=True,fg=DARK,bg=bg,h='right')
        C(ws,r,3,rec.get('fees',''),bg=bg); C(ws,r,4,rec.get('request_date',''),bg=bg)
        C(ws,r,5,rec.get('client_reason',''),bg=bg,h='right',wrap=True)
        C(ws,r,6,rec.get('sales_notes',''),bg=bg,h='right',wrap=True)
        tc=ws.cell(r,7,t)
        tc.font=Font(name='Arial',bold=True,size=10,color=type_colors.get(t,'000000'))
        tc.fill=PatternFill('solid',start_color=bg); tc.alignment=Al(); tc.border=bd()
        C(ws,r,8,rec.get('week_label',''),bg=bg); RH(ws,r,20)
    for c,w in [(1,5),(2,22),(3,14),(4,14),(5,30),(6,34),(7,12),(8,14)]: CW(ws,c,w)
    ws.freeze_panes='A4'
    
    # لوحة التحليل
    la=wb.create_sheet('لوحة التحليل'); la.sheet_view.rightToLeft=True
    MC(la,1,1,8,f'لوحة تحليل عملاء الاسترداد  |  نورة لتعليم القيادة',size=13); RH(la,1,34)
    KPI(la,3,1,'إجمالي العملاء',total,DARK)
    KPI(la,4,1,'طلبات الاسترداد',sum(1 for r in recs if r.get('request_type')=='استرداد'),RED)
    KPI(la,5,1,'طلبات الاستكمال',sum(1 for r in recs if r.get('request_type')=='استكمال'),'006B3C')
    KPI(la,6,1,'طلبات التأجيل',sum(1 for r in recs if r.get('request_type')=='تأجيل'),ORANGE)
    total_fees=sum(r.get('fees') or 0 for r in recs if r.get('request_type')=='استرداد')
    KPI(la,7,1,'إجمالي المبالغ المطلوبة',f'{total_fees:,.0f} ريال',TEAL2)
    for c,w in enumerate([22,10,10,10,22,10,10,10],1): CW(la,c,w)
    return wb

def build_ratings(recs):
    wb=Workbook()
    ws=wb.active; ws.title='نتائج التقييم'; ws.sheet_view.rightToLeft=True
    total=len(recs)
    MC(ws,1,1,6,f'تقرير الأعلى تقييماً – مبني على آراء العملاء  |  نورة لتعليم القيادة',size=13); RH(ws,1,34)
    MC(ws,2,1,6,f'إجمالي المدربين المذكورين: {total}',bold=False,bg=TEAL,size=10); RH(ws,2,22)
    TH(ws,3,['الترتيب','اسم المدرب/ة','عدد مرات الذكر','النسبة%','التصنيف','تاريخ الرفع'])
    medals=['🥇','🥈','🥉']
    medal_bgs=['FFF8DC','F5F5EE','F0EEE8']
    for ri,rec in enumerate(recs):
        r=4+ri; bg=medal_bgs[ri] if ri<3 else (ALT1 if ri%2==0 else ALT2)
        rank=rec.get('rank_position',ri+1)
        medal=medals[ri] if ri<3 else f'#{rank}'
        mc=ws.cell(r,1,medal)
        mc.font=Font(name='Arial',bold=True,size=12); mc.fill=PatternFill('solid',start_color=bg); mc.alignment=Al(); mc.border=bd()
        C(ws,r,2,rec.get('trainer_name',''),bold=(ri<3),fg=DARK,bg=bg,h='right')
        C(ws,r,3,rec.get('mention_count',''),bg=bg)
        C(ws,r,4,f"{rec.get('mention_pct',0):.1f}%" if rec.get('mention_pct') else '-',bg=bg)
        tag='أعلى تقييماً' if ri==0 else 'ثاني أعلى' if ri==1 else 'ثالث أعلى' if ri==2 else 'ضمن أعلى 10' if ri<10 else ''
        C(ws,r,5,tag,bold=(ri<3),fg='006B3C' if ri<3 else '000000',bg=bg)
        C(ws,r,6,rec.get('upload_date',''),bg=bg); RH(ws,r,22)
    for c,w in [(1,8),(2,28),(3,14),(4,12),(5,18),(6,14)]: CW(ws,c,w)
    ws.freeze_panes='A4'
    return wb

# ═══════════════════════════════════════════
# VERCEL HANDLER
# ═══════════════════════════════════════════
def handler(request):
        
    # CORS headers
    headers = {
        'Access-Control-Allow-Origin': '*',
        'Access-Control-Allow-Methods': 'POST, OPTIONS',
        'Access-Control-Allow-Headers': 'Content-Type',
    }
    
    if request.method == 'OPTIONS':
        return Response('', 200, headers)
    
    if request.method != 'POST':
        return Response('Method not allowed', 405, headers)
    
    try:
        body = json.loads(request.body)
        rtype = body.get('type', '')
        recs = body.get('data', [])
        year = body.get('year', 2026)
        month = body.get('month', 1)
        section_name = body.get('section_name', 'المخالفات')

        builders = {
            'violations': lambda: build_violations(recs, section_name, year),
            'surveys':    lambda: build_surveys(recs, year),
            'facility':   lambda: build_facility(recs, year),
            'building':   lambda: build_building(recs, year),
            'attendance': lambda: build_attendance(recs, month, year),
            'incentives': lambda: build_incentives(recs, year),
            'ops':        lambda: build_ops(recs, year),
            'sales':      lambda: build_sales(recs, year),
            'ratings':    lambda: build_ratings(recs),
        }

        if rtype not in builders:
            return Response('Unknown type', 400, headers)

        wb = builders[rtype]()
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        xlsx_bytes = buf.read()

        resp_headers = {
            **headers,
            'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'Content-Disposition': f'attachment; filename="report.xlsx"',
        }
        return Response(xlsx_bytes, 200, resp_headers)

    except Exception as e:
        return Response(str(e), 500, headers)


class Response:
    def __init__(self, body, status=200, headers=None):
        self.body = body
        self.status_code = status
        self.headers = headers or {}
